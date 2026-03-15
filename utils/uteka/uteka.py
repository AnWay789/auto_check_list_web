import csv
import gzip
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import httpx
from openpyxl import load_workbook

from config.celery import app
from config.settings import (
    BASE_DIR,
    UTEKA_API_KEY,
    UTEKA_BASE_URL,
    UTEKA_SHARES_XLSX_URL,
)

# Каталог для накопления выгрузок Ютека (price_*.csv, share_*.csv)
UTEKA_EXPORT_DIR = BASE_DIR / "data" / "uteka"

OUTPUT_COLUMNS = [
    'altProductId',
    'priceType',
    'Здравсити',
    'Apteka.ru',
    'Доктор Столетов',
    'Аптеки ГОРЗДРАВ',
    'Планета Здоровья',
    'Ваша №1',
    'ЕАПТЕКА',
]


def get_uteka_price_data(
    output_path: str | Path,
    raw_output_path: str | Path | None = None,
) -> list[dict]:
    """
    Скачивает выгрузку цен Ютека, оставляет только priceType=median,
    сохраняет в CSV только altProductId, priceType и цены семи конкурентов.
    Если передан raw_output_path — туда сохраняется необработанный CSV (распакованный gzip).
    """
    url = f"{UTEKA_BASE_URL}/export/partner-prices/{UTEKA_API_KEY}/partner-prices-puls-city-msk.csv.gz"
    response = httpx.get(url)
    response.raise_for_status()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    raw_content = gzip.decompress(response.content).decode("utf-8")

    if raw_output_path is not None:
        Path(raw_output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(raw_output_path, "w", encoding="utf-8") as f:
            f.write(raw_content)

    rows_filtered: list[dict] = []
    reader = csv.DictReader(raw_content.splitlines())
    for row in reader:
        if row.get("priceType") != "median":
            continue
        out_row = {k: row.get(k, "") for k in OUTPUT_COLUMNS}
        rows_filtered.append(out_row)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows_filtered)

    return rows_filtered


def _parse_excel_date(cell_value):
    """Приводит значение ячейки к date для сравнения с сегодня."""
    if cell_value is None:
        return None
    if isinstance(cell_value, date) and not isinstance(cell_value, datetime):
        return cell_value
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, str):
        try:
            return datetime.strptime(cell_value.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _parse_percent_value(value):
    """Парсит число из ячейки (может быть float или строка с запятой)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _cell_to_csv_value(value):
    """Преобразует значение ячейки в строку для CSV."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        return str(value) if value == value else ""  # avoid NaN
    return str(value)


def get_uteka_share_data(
    output_path: str | Path,
    raw_output_path: str | Path | None = None,
) -> list[dict]:
    """
    Скачивает выгрузку долей по регионам Ютека (xlsx), оставляет только строки,
    где в колонке Date указана текущая дата. Сохраняет в CSV все столбцы (Date, City,
    AllTypesPercent, ExtendedPickupPercent) без агрегации.
    Если передан raw_output_path — туда сохраняется исходный xlsx без обработки.
    """
    response = httpx.get(UTEKA_SHARES_XLSX_URL)
    response.raise_for_status()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if raw_output_path is not None:
        Path(raw_output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(raw_output_path, "wb") as f:
            f.write(response.content)

    wb = load_workbook(BytesIO(response.content), read_only=True)
    ws = wb.active
    if not ws:
        wb.close()
        raise ValueError("В xlsx нет активного листа")

    today = date.today()
    rows_iter = ws.iter_rows(min_row=1, max_col=4, values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        raise ValueError("В xlsx нет строки заголовков")

    # Заголовки: используем как есть или подставляем дефолтные
    fieldnames = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row[:4])]
    col_date_index = 0  # первая колонка — дата

    rows_today: list[list[str]] = []
    for row in rows_iter:
        if len(row) < 4:
            continue
        row_date = _parse_excel_date(row[col_date_index])
        if row_date != today:
            continue
        rows_today.append([_cell_to_csv_value(v) for v in row[:4]])

    wb.close()

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        writer.writerows(rows_today)

    return [dict(zip(fieldnames, r)) for r in rows_today]


def _uteka_export_filename(prefix: str, suffix: str = "csv") -> str:
    """Имя файла с датой и временем: prefix_DD_MM_YYYY_HH_MM.suffix"""
    now = datetime.now()
    return f"{prefix}_{now:%d_%m_%Y_%H_%M}.{suffix}"


@app.task(
    name="utils.uteka.run_uteka_price_task",
    max_retries=3,
    autoretry_for=(Exception,),
)
def run_uteka_price_task() -> list[dict]:
    """
    Периодическая задача: выгрузка цен конкурентов Ютека (Москва и МО).
    Сохраняет обработанный файл price_DD_MM_YYYY_HH_MM.csv и сырой raw_price_DD_MM_YYYY_HH_MM.csv.
    """
    UTEKA_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = UTEKA_EXPORT_DIR / _uteka_export_filename("price")
    raw_path = UTEKA_EXPORT_DIR / _uteka_export_filename("raw_price")
    return get_uteka_price_data(path, raw_output_path=raw_path)


@app.task(
    name="utils.uteka.run_uteka_share_task",
    max_retries=3,
    autoretry_for=(Exception,),
)
def run_uteka_share_task() -> list[dict]:
    """
    Периодическая задача: выгрузка долей по регионам Ютека за текущий день.
    Сохраняет обработанный файл share_DD_MM_YYYY_HH_MM.csv и сырой raw_share_DD_MM_YYYY_HH_MM.xlsx.
    """
    UTEKA_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = UTEKA_EXPORT_DIR / _uteka_export_filename("share")
    raw_path = UTEKA_EXPORT_DIR / _uteka_export_filename("raw_share", suffix="xlsx")
    return get_uteka_share_data(path, raw_output_path=raw_path)
